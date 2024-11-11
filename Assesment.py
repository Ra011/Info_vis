import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
import time
from openpyxl import load_workbook

base_file_path = "C:/Users/Student/OneDrive/Documents/Docs2/uni/year 3/semester 1/COMP3736 Information Visualization/cw/data/transposed_seasonal_absence_data_"
file_extension = ".xlsx"
existing_file_path = "C:/Users/Student/OneDrive/Documents/Docs2/uni/year 3/semester 1/COMP3736 Information Visualization/cw/data/response_data.xlsx"

response_data_scatter = []
response_data_heatmap = []

# Initial user selection
print("Please choose the type of graph to display first:")
print("1. Scatter Plots")
print("2. Heat Maps")
while True:
    graph_choice = input("Enter 1 for Scatter Plots first or 2 for Heat Maps first: ")
    if graph_choice in ["1", "2"]:
        break
    else:
        print("Invalid input. Please enter 1 or 2.")

# Display welcome message and confirm understanding
print("""
Welcome to the experiment, and thank you for helping us out!

In this experiment, you will be shown a series of graphs. Each graph displays the average attendance in 10 different schools across the Autumn, Spring, and Summer semesters.

For each graph, you will be asked to identify which semester has the highest average attendance based on your observation. 

To enter your response, type:
- '1' for Autumn
- '2' for Spring
- '3' for Summer
""")

while True:
    user_confirmation = input("Please type '1' to confirm that you understand and are ready to begin: ")
    if user_confirmation == "1":
        break
    else:
        print("Invalid input. Please enter 1 to confirm.")

# Functions for displaying scatter plot and heat map
def display_scatter(graph_data, graph_number):
    semesters = ['Autumn', 'Spring', 'Summer']
    values = graph_data[semesters].values.flatten()
    semester_labels = [semester for semester in semesters for _ in range(len(graph_data))]

    plt.figure(figsize=(8, 6))
    plt.scatter(semester_labels, values, color='blue')
    plt.xlabel("Semester")
    plt.ylabel("Attendance %")
    plt.title(f"Graph {graph_number}: Attendance by Semester")
    plt.show(block=False)

    start_time = time.time()
    while True:
        user_input = input("Enter your response (1, 2, or 3): ")
        if user_input in ["1", "2", "3"]:
            break
        else:
            print("Invalid input. Please enter 1, 2, or 3.")

    response_time = time.time() - start_time
    response_data_scatter.append([graph_number, user_input, response_time])

    plt.close()
    blank_screen()

def display_heatmap(graph_data, graph_number):
    semesters = ['Autumn', 'Spring', 'Summer']
    data = graph_data[semesters].values

    fig, ax = plt.subplots(figsize=(8, 6))
    cax = ax.imshow(data, cmap="YlGnBu", aspect="auto")
    plt.colorbar(cax)
    ax.set_xticks(np.arange(len(semesters)))
    ax.set_xticklabels(semesters, fontsize=11)
    ax.set_yticks(np.arange(len(graph_data)))
    ax.set_yticklabels(graph_data.index, fontsize=11)
    plt.xlabel("Semester", fontsize=11)
    plt.ylabel("School", fontsize=11)
    plt.title(f"Heat Map {graph_number}: Attendance by Semester", fontsize=13)
    plt.show(block=False)

    start_time = time.time()
    while True:
        user_input = input("Enter your response (1, 2, or 3): ")
        if user_input in ["1", "2", "3"]:
            break
        else:
            print("Invalid input. Please enter 1, 2, or 3.")

    response_time = time.time() - start_time
    response_data_heatmap.append([graph_number, user_input, response_time])

    plt.close()
    blank_screen()

def blank_screen():
    plt.figure(figsize=(8, 6))
    plt.axis('off')
    plt.show(block=False)
    time.sleep(1)
    plt.close()

# Run the experiment in the specified alternating order
for i in range(1, 11):
    df = pd.read_excel(f"{base_file_path}{i}{file_extension}")
    if graph_choice == "1":
        display_scatter(df, i)
        display_heatmap(df, i)
    else:
        display_heatmap(df, i)
        display_scatter(df, i)

# Write results to Excel file
try:
    workbook = load_workbook(existing_file_path)
    sheet = workbook.active

    # Add scatter plot responses to first table
    next_row = sheet.max_row + 2 if sheet.max_row > 1 else 1
    for entry in response_data_scatter:
        sheet.append(entry)

    heatmap_start_col = 7
    for idx, entry in enumerate(response_data_heatmap, start=next_row):
        for j, value in enumerate(entry):
            sheet.cell(row=idx, column=heatmap_start_col + j, value=value)

    workbook.save(existing_file_path)
    print("Responses saved.")

except FileNotFoundError:
    response_df_scatter = pd.DataFrame(response_data_scatter, columns=["Graph", "Response", "Response Time"])
    response_df_heatmap = pd.DataFrame(response_data_heatmap, columns=["Graph", "Response", "Response Time"])
    with pd.ExcelWriter(existing_file_path) as writer:
        response_df_scatter.to_excel(writer, index=False, sheet_name="Scatter Data")
        response_df_heatmap.to_excel(writer, startcol=heatmap_start_col, index=False, sheet_name="Heatmap Data")
    print("File not found. A new file has been created.")
